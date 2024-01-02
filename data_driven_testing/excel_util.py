# module contains important functions
# to perform data read/write operations

import openpyxl
from openpyxl.styles import PatternFill


def get_row_count(file, sheetName):
    # workbook = openpyxl.load_workbook(file)
    # sheet = workbook[sheetName]
    sheet = openpyxl.load_workbook(file)[sheetName]
    return sheet.max_row


def get_col_count(file, sheetName):
    # workbook = openpyxl.load_workbook(file)
    # sheet = workbook[sheetName]
    sheet = openpyxl.load_workbook(file)[sheetName]
    return sheet.max_column


def read_data(file, sheetName, rowNo, columnNo):
    # workbook = openpyxl.load_workbook(file)
    # sheet = workbook[sheetName]
    sheet = openpyxl.load_workbook(file)[sheetName]

    # return sheet.cell(rowNo, columnNo).value
    return openpyxl.load_workbook(file)[sheetName].cell(rowNo, columnNo).value


def write_data(file, sheetName, rowNo, columnNo, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(rowNo, columnNo).value = data
    workbook.save(file)
