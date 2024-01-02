import openpyxl
import os
import pdb

cwd = os.getcwd()
# print('Current working directory: '+cwd)
file = cwd+'\\excel_demo_db.xlsx'
print(file)

# # load the workbook > sheet
# workbook = openpyxl.load_workbook(file)
# sheet1 = workbook["Sheet1"]

# or it can be written as
# sheet = openpyxl.load_workbook(file)['Sheet1']

# or to select active sheet
sheet = openpyxl.load_workbook(file).active

max_row = sheet.max_row
max_col = sheet.max_column

for row in range(1, max_row+1):
    for col in range(1, max_col):
        print(sheet.cell(row, col).value)
    print()
