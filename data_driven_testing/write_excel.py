import openpyxl

file = "D:\\pyc_project\\Selenium\\data_driven_testing\\excel_demo_db.xlsx"
# print(file)

workbook = openpyxl.load_workbook(file)
sheet = workbook['w_sheet']

sheet.cell(1, 1).value = 'Book'
sheet.cell(1, 2).value = 'QTY'
sheet.cell(1, 3).value = 'Amount'

sheet.cell(2, 1).value = 'Java'
sheet.cell(2, 2).value = 1
sheet.cell(2, 3).value = 259

sheet.cell(3, 1).value = 'Python'
sheet.cell(3, 2).value = 2
sheet.cell(3, 3).value = 560

workbook.save(file)
